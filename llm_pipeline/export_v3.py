"""
Export v3 - 从 projects_v3 表导出 15 字段商业情报 Excel

用法:
    python llm_pipeline/export_v3.py

依赖: openpyxl
输出: ~/Desktop/体育招投标_商业情报单.xlsx
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from llm_pipeline.db import get_conn, DEFAULT_DB_PATH

# ─── 常量 ───────────────────────────────

HEADERS = [
    "序号", "项目名称", "时效层级", "项目状态", "优先级", "标记理由",
    "预算/成交金额", "供应商", "投标截止时间", "采购单位", "公告类型",
    "发布时间", "项目编号", "原文链接", "项目描述",
]

DEFAULT_OUTPUT = os.path.join(
    os.path.expanduser("~"), "Desktop", "体育招投标_商业情报单.xlsx"
)

# 样式
H_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
H_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

D_FONT = Font(name="微软雅黑", size=10)
D_ALIGN = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
Y_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

COL_WIDTHS = {
    1: 6, 2: 60, 3: 15, 4: 15, 5: 12, 6: 25, 7: 20, 8: 25,
    9: 20, 10: 20, 11: 15, 12: 15, 13: 20, 14: 40, 15: 40,
}

FIELD_MAP = [
    "id", "title", "timeliness_stratum", "project_status", "priority_tag",
    "priority_reason", "budget_or_winning", "supplier", "submission_deadline",
    "buyer", "notice_type", "publish_date", "project_number", "source_url",
    "scope_description",
]


def record_to_row(record):
    """将 projects_v3 行记录转换为 15 列列表"""
    row = []
    for field in FIELD_MAP:
        val = record.get(field)
        if val is None or val == "":
            val = "原网页未披露该项"
        row.append(val)
    return row


def apply_headers(ws):
    """写入表头样式"""
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = H_FONT
        c.fill = H_FILL
        c.alignment = H_ALIGN
        c.border = THIN_BORDER


def apply_column_widths(ws):
    """设置列宽"""
    for ci, w in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(ci)].width = w


def apply_auto_filter(ws):
    """启用自动筛选"""
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"


def export_to_excel(db_path=None, output_path=None):
    """主导出函数"""
    conn = get_conn(db_path)
    output = output_path or DEFAULT_OUTPUT

    # 读取所有记录
    rows = conn.execute("""
        SELECT * FROM projects_v3
        ORDER BY publish_date DESC, id DESC
    """).fetchall()
    records = [dict(r) for r in rows]

    if not records:
        print("projects_v3 表无数据，跳过导出")
        conn.close()
        return

    wb = Workbook()

    # ── Sheet 1: 完整数据 ──
    ws1 = wb.active
    ws1.title = "完整数据"
    apply_headers(ws1)

    for ri, rec in enumerate(records, 2):
        vals = record_to_row(rec)
        is_yes = (rec.get("priority_tag") or "") == "YES"
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(row=ri, column=ci, value=val)
            c.font = D_FONT
            c.alignment = D_ALIGN
            c.border = THIN_BORDER
            if is_yes:
                c.fill = Y_FILL

    ws1.freeze_panes = "A2"
    apply_column_widths(ws1)
    apply_auto_filter(ws1)

    # ── Sheet 2: 高价值汇总 ──
    ws2 = wb.create_sheet("高价值汇总")
    apply_headers(ws2)

    row_idx = 2
    for rec in records:
        if (rec.get("priority_tag") or "") != "YES":
            continue
        vals = record_to_row(rec)
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=row_idx, column=ci, value=val)
            c.font = D_FONT
            c.alignment = D_ALIGN
            c.border = THIN_BORDER
            c.fill = Y_FILL
        row_idx += 1

    ws2.freeze_panes = "A2"
    apply_column_widths(ws2)
    apply_auto_filter(ws2)

    # 保存
    os.makedirs(os.path.dirname(output), exist_ok=True)
    wb.save(output)

    conn.close()

    print(f"  Sheet1: {ws1.max_row - 1} records")
    print(f"  Sheet2: {ws2.max_row - 1} records (YES only)")
    print(f"  Output: {output}")
    return output


def main():
    print("=" * 50)
    print("Export v3 - projects_v3 -> Excel")
    print("=" * 50)
    db = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "llm_bidding_v2.db"
    )
    out = export_to_excel(db_path=db)
    print(f"\nExport complete: {out}")


if __name__ == "__main__":
    main()
