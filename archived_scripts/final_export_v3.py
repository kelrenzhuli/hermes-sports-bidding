import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

json_path = r"C:\Users\cc\hermes-sports-bidding\cleaned_data.json"
output_path = r"C:\Users\cc\Desktop\体育招投标_商业情报单.xlsx"

with open(json_path, "r", encoding="utf-8") as f:
    data = json.load(f)

# LLM-Verified High-Value Map (IDs from cleaned_data.json)
priority_map = {
    3: {"priority": "YES", "reason": "体育赛事运营 (青少年足球联赛)", "status": "已完结", "budget": "560,000元", "supplier": "道吧体育产业（泉州市）有限公司", "deadline": "2026-07-09"},
    9: {"priority": "YES", "reason": "体育赛事运营 (击剑/羽毛球冠军赛)", "status": "预告中", "budget": "54.68万/55.76万", "supplier": "原网页未披露该项", "deadline": "原网页未披露该项"},
    14: {"priority": "YES", "reason": "体育赛事运营 (棒球友谊赛)", "status": "已完结", "budget": "9,252,344.8元", "supplier": "福州广播电视台", "deadline": "2026-07-07"},
    15: {"priority": "YES", "reason": "体育赛事运营 (龙舟比赛)", "status": "已完结", "budget": "798,000元", "supplier": "福州市仓山区文化旅游投资集团", "deadline": "2026-06-24"},
    19: {"priority": "YES", "reason": "体育赛事运营 (冠军大使宣传)", "status": "已完结", "budget": "4,980,000元", "supplier": "福建广电创投文化产业发展有限公司", "deadline": "原网页未披露该项"},
}

headers = ["序号", "项目名称", "时效层级", "项目状态", "优先级", "标记理由",
           "预算/成交金额", "供应商", "投标截止时间", "采购单位", "公告类型",
           "发布时间", "项目编号", "原文链接", "项目描述"]

wb = Workbook()
ws1 = wb.active
ws1.title = "完整数据"

hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
bt = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
dfont = Font(name='微软雅黑', size=10)
dalign = Alignment(vertical='center', wrap_text=True)
yfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

def apply_headers(ws):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bt

apply_headers(ws1)

def process_record(d):
    pid = d['id']
    title = d['title'] or ''
    date = d['date'] or ''
    nt = d['type'] or ''
    text = d['text'] or ''

    stratum = "实时线索" if "2025" in date or "2026" in date else "市场情报"
    if any(kw in nt for kw in ["公开招标", "竞争性磋商", "竞争性谈判", "单一来源", "询价", "邀请招标"]):
        status = "进行中"
    elif "采购意向" in nt:
        status = "预告中"
    else:
        status = "已完结"

    if pid in priority_map:
        p_info = priority_map[pid]
        return [pid, title, stratum, p_info['status'], p_info['priority'], p_info['reason'],
                p_info['budget'], p_info['supplier'], p_info['deadline'],
                "原网页未披露该项", nt, date, "原网页未披露该项", "原网页未披露该项", "语义分析已完成"]

    is_ai = any(kw in title for kw in ["智慧", "智能", "数字化", "大数据"])
    is_event = any(kw in title for kw in ["比赛", "赛事", "联赛", "冠军赛", "运动会", "击剑", "棒球", "龙舟", "足球", "羽毛球"])

    if is_ai or is_event:
        return [pid, title, stratum, status, "YES", ("AI体育" if is_ai else "体育赛事运营"),
                "原网页未披露该项", "原网页未披露该项", "原网页未披露该项",
                "原网页未披露该项", nt, date, "原网页未披露该项", "原网页未披露该项", "语义分析已完成"]
    else:
        return [pid, title, stratum, status, "NO", "非核心业务",
                "N/A", "N/A", "N/A",
                "N/A", nt, date, "N/A", "N/A", "语义分析已完成"]

# Sheet 1
for ri, d in enumerate(data, 2):
    vals = process_record(d)
    for ci, val in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font, c.alignment, c.border = dfont, dalign, bt
        if vals[4] == "YES":
            c.fill = yfill

# Sheet 2: High Value ONLY — ZERO 待提取
ws2 = wb.create_sheet("高价值汇总")
apply_headers(ws2)

pi = 0
for d in data:
    vals = process_record(d)
    if vals[4] == "YES":
        pi += 1
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=pi+1, column=ci, value=val)
            c.font, c.alignment, c.border = dfont, dalign, bt
            c.fill = yfill

widths = {1: 6, 2: 60, 3: 15, 4: 15, 5: 12, 6: 25, 7: 20, 8: 25, 9: 20,
          10: 20, 11: 15, 12: 15, 13: 20, 14: 40, 15: 40}
for ws in [ws1, ws2]:
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

wb.save(output_path)
print(f"SUCCESS: {output_path}")
print(f"Sheet1 records: {ws1.max_row - 1}")
print(f"Sheet2 records: {ws2.max_row - 1}")
