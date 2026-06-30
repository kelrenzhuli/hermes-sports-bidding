# -*- coding: utf-8 -*-
"""从v2数据库导出已爬取的245条数据到桌面Excel"""
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

db = r'C:\Users\cc\hermes-sports-bidding\llm_bidding_v2.db'
output = r'C:\Users\cc\Desktop\体育招投标_原始数据.xlsx'

conn = sqlite3.connect(db)
conn.row_factory = sqlite3.Row
rows = conn.execute("""
    SELECT id, title, url, notice_type, region, publish_date, project_number, procurement_method
    FROM pages
    ORDER BY id
""").fetchall()
conn.close()

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "爬取数据"

# 样式
hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
dfont = Font(name='微软雅黑', size=10)
dalign = Alignment(vertical='center', wrap_text=True)
ufont = Font(name='微软雅黑', size=10, color='0563C1', underline='single')
bt = Border(left=Side(style='thin', color='D0D0D0'),right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),bottom=Side(style='thin', color='D0D0D0'))

bidding_types = {"公开招标","竞争性磋商","竞争性谈判","单一来源","询价","邀请招标"}
ev_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
bid_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

headers = ["序号","项目名称","原文链接","公告类型","采购方式","地区","发布时间","项目编号"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bt

for ri, row in enumerate(rows, 2):
    nt = row['notice_type'] or ''
    is_ev = nt in bidding_types
    fill = ev_fill if is_ev else bid_fill if nt in ("合同","结果") else None
    
    vals = [row['id'], row['title'], row['url'] or '', nt, row['procurement_method'] or '',
            row['region'] or '', row['publish_date'] or '', row['project_number'] or '']
    
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.font = dfont
        c.alignment = dalign
        c.border = bt
        if fill: c.fill = fill
        if ci == 3 and str(val).startswith('http'):
            c.font = ufont
            c.hyperlink = val

# 列宽
widths = {1:6, 2:65, 3:50, 4:14, 5:16, 6:16, 7:16, 8:30}
for ci, w in widths.items():
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:H{len(rows)+1}"

wb.save(output)

# 统计
bidding = sum(1 for r in rows if r['notice_type'] in bidding_types)
contracts = sum(1 for r in rows if r['notice_type'] == '合同')
intents = sum(1 for r in rows if r['notice_type'] == '采购意向')
has_proj = sum(1 for r in rows if r['project_number'])
has_method = sum(1 for r in rows if r['procurement_method'])

print(f"📁 导出完成: {output}")
print(f"  总记录: {len(rows)}")
print(f"  招标类: {bidding}")
print(f"  合同类: {contracts}")
print(f"  采购意向: {intents}")
print(f"  有项目编号: {has_proj}")
print(f"  有采购方式: {has_method}")
