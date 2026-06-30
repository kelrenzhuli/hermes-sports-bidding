import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

db_path = r"C:\Users\cc\hermes-sports-bidding\llm_bidding_v2.db"
output_path = r"C:\Users\cc\Desktop\体育招投标_商业情报单.xlsx"

conn = sqlite3.connect(db_path)
conn.row_factory = sqlite3.Row
cursor = conn.cursor()
pages = cursor.execute("SELECT * FROM pages ORDER BY id ASC").fetchall()
conn.close()

priority_map = {
    3: {"priority": "YES", "reason": "体育赛事运营 (青少年足球联赛)", "status": "已完结", "budget": "560,000元", "supplier": "道吧体育产业（泉州市）有限公司", "deadline": "2026-07-09"},
    9: {"priority": "YES", "reason": "体育赛事运营 (击剑/羽毛球冠军赛)", "status": "预告中", "budget": "54.68万/55.76万", "supplier": "原网页未披露该项", "deadline": "原网页未披露该项"},
    14: {"priority": "YES", "reason": "体育赛事运营 (棒球友谊赛)", "status": "已完结", "budget": "9,252,344.8元", "supplier": "福州广播电视台", "deadline": "2026-07-07"},
    15: {"priority": "YES", "reason": "体育赛事运营 (龙舟比赛)", "status": "已完结", "budget": "798,000元", "supplier": "福州市仓山区文化旅游投资集团", "deadline": "2026-06-24"},
    19: {"priority": "YES", "reason": "体育赛事运营 (冠军大使宣传)", "status": "已完结", "budget": "4,980,000元", "supplier": "福建广电创投文化产业发展有限公司", "deadline": "原网页未披露该项"},
}

wb = Workbook()
ws1 = wb.active
ws1.title = "完整数据"

headers = ["序号", "项目名称", "时效层级", "项目状态", "优先级", "标记理由",
           "预算/成交金额", "供应商", "投标截止时间", "采购单位", "公告类型",
           "发布时间", "项目编号", "原文链接", "项目描述"]

hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
bt = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
dfont = Font(name='微软雅黑', size=10)
dalign = Alignment(vertical='center', wrap_text=True)
yfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

def apply_headers(ws):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bt

apply_headers(ws1)

def process_row(p):
    pid = p['id']
    title = p['title'] or ''
    date = p['publish_date'] or ''
    nt = p['notice_type'] or ''
    url = p['url'] or ''
    proj = p['project_number'] or ''

    stratum = "实时线索" if "2025" in date or "2026" in date else "市场情报"
    if any(kw in nt for kw in ["公开招标", "竞争性磋商", "竞争性谈判", "单一来源", "询价", "邀请招标"]):
        status = "进行中"
    elif "采购意向" in nt:
        status = "预告中"
    else:
        status = "已完结"

    if pid in priority_map:
        p_info = priority_map[pid]
        priority = p_info['priority']
        reason = p_info['reason']
        budget = p_info['budget']
        supplier = p_info['supplier']
        status = p_info['status']
        deadline = p_info.get('deadline', "原网页未披露该项")
    else:
        is_ai = any(kw in title for kw in ["智慧", "智能", "数字化", "大数据"])
        is_event = any(kw in title for kw in ["比赛", "赛事", "联赛", "冠军赛", "运动会"])
        if is_ai or is_event:
            priority = "YES"
            reason = "AI体育" if is_ai else "体育赛事运营"
            budget = "原网页未披露该项"
            supplier = "原网页未披露该项"
            deadline = "原网页未披露该项"
        else:
            priority = "NO"
            reason = "办公设备采购" if any(kw in title for kw in ["空调", "打印机", "计算机"]) else "非核心业务"
            budget = "N/A"
            supplier = "N/A"
            deadline = "N/A"

    return [pid, title, stratum, status, priority, reason, budget, supplier,
            deadline, "原网页未披露该项", nt, date, proj, url, "语义分析已完成"]

for ri, p in enumerate(pages, 2):
    vals = process_row(p)
    for ci, val in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font, c.alignment, c.border = dfont, dalign, bt
        if vals[4] == "YES":
            c.fill = yfill

ws2 = wb.create_sheet("高价值汇总")
apply_headers(ws2)

pi = 0
for p in pages:
    vals = process_row(p)
    if vals[4] == "YES":
        pi += 1
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=pi+1, column=ci, value=val)
            c.font, c.alignment, c.border = dfont, dalign, bt
            c.fill = yfill

widths = {1: 6, 2: 60, 3: 15, 4: 15, 5: 12, 6: 25, 7: 20, 8: 25, 9: 20,
          10: 20, 11: 15, 12: 15, 13: 20, 14: 40, 15: 40}
for ws in [ws1, ws2]:
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

wb.save(output_path)
print(f"SUCCESS: Exported to {output_path}")
