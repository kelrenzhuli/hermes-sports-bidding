# -*- coding: utf-8 -*-
"""将数据库全部清洗结果导出为Excel到桌面"""
import sqlite3, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

db = r'C:\Users\cc\hermes-sports-bidding\llm_bidding_v2.db'
output = r'C:\Users\cc\Desktop\体育招投标_清洗数据总表_每日更新.xlsx'

conn = sqlite3.connect(db)
conn.row_factory = sqlite3.Row

# 获取所有记录（page + extracted 联表）
rows = conn.execute("""
    SELECT p.id, p.title, p.url, p.source, p.notice_type, p.region, 
           p.publish_date, p.crawled_at, p.status,
           e.data_json
    FROM pages p
    JOIN extracted e ON p.id = e.page_id
    ORDER BY 
        CASE WHEN json_extract(e.data_json, '$.priority_tag') = 'YES' THEN 0 ELSE 1 END,
        p.id
""").fetchall()

conn.close()

# 创建 Excel
wb = Workbook()
ws = wb.active
ws.title = "完整数据"

# === 样式定义 ===
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

priority_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # 浅黄
normal_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

data_font = Font(name='微软雅黑', size=10)
data_align = Alignment(vertical='center', wrap_text=True)
url_font = Font(name='微软雅黑', size=10, color='0563C1', underline='single')

# === 表头 ===
headers = [
    "序号", "项目名称", "原文链接", "公告类型", "采购方式", "所属行业",
    "地区", "信息来源", "发布时间", "预算金额", "合同金额（中标价）", "中标供应商",
    "中标金额", "采购单位", "联系方式", "资质要求",
    "投标截止时间", "招标范围/采购内容", "中标/变更内容",
    "重点标记", "标记理由", "爬取时间"
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# === 数据行 ===
idx = 0
for row in rows:
    idx += 1
    pid = row['id']
    title = row['title']
    url = row['url'] or ""
    source = row['source'] or ""
    nt = row['notice_type'] or ""
    region = row['region'] or ""
    pub_date = row['publish_date'] or ""
    crawled = row['crawled_at'] or ""
    
    # 解析 JSON
    data = json.loads(row['data_json']) if row['data_json'] else {}
    
    priority_tag = data.get('priority_tag', 'NO')
    priority_reason = data.get('priority_reason', '')
    industry = data.get('industry', '') or ''
    budget = data.get('budget_amount', '') or ''
    contract_amount = data.get('contract_amount', '') or ''
    procurement = data.get('procurement_method', '') or ''
    winner = data.get('winner', '') or ''
    winning_price = data.get('winning_price', '') or ''
    buyer = data.get('buyer', '') or ''
    contact = data.get('contact', '') or ''
    eligibility = data.get('eligibility', '') or ''
    submission_deadline = data.get('submission_deadline', '') or ''
    scope = data.get('scope_of_work', '') or ''
    amendment = data.get('amendment', '') or ''
    
    values = [
        pid, title, url, nt, procurement, industry,
        region, source, pub_date, budget, contract_amount, winner,
        winning_price, buyer, contact, eligibility,
        submission_deadline, scope, amendment,
        priority_tag, priority_reason, crawled
    ]
    
    row_num = idx + 1
    fill = priority_fill if priority_tag == 'YES' else normal_fill
    
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        cell.fill = fill
        
        # 链接列特殊处理
        if col_idx == 3 and val and val.startswith('http'):
            cell.font = url_font
            cell.hyperlink = val

# === 列宽 ===
col_widths = {
    1: 6,    # 序号
    2: 55,   # 项目名称
    3: 40,   # 原文链接
    4: 14,   # 公告类型
    5: 14,   # 采购方式
    6: 12,   # 所属行业
    7: 16,   # 地区
    8: 18,   # 信息来源
    9: 14,   # 发布时间
    10: 14,  # 预算金额
    11: 16,  # 合同金额
    12: 22,  # 中标供应商
    13: 16,  # 中标金额
    14: 22,  # 采购单位
    15: 20,  # 联系方式
    16: 25,  # 资质要求
    17: 16,  # 投标截止
    18: 35,  # 招标范围
    19: 20,  # 变更内容
    20: 10,  # 重点标记
    21: 25,  # 标记理由
    22: 16,  # 爬取时间
}

for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# 冻结首行
ws.freeze_panes = 'A2'

# 自动筛选
ws.auto_filter.ref = f"A1:V{idx+1}"

# === 第二个 Sheet: 重点项目汇总 ===
ws2 = wb.create_sheet("重点项目")
ws2_headers = [
    "序号", "项目名称", "标签理由", "预算金额", "中标金额",
    "采购单位", "中标供应商", "地区", "公告类型", "原文链接"
]

# 表头
for col_idx, h in enumerate(ws2_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    cell.alignment = header_align
    cell.border = thin_border

# 数据 - 重新打开连接
conn2 = sqlite3.connect(db)
conn2.row_factory = sqlite3.Row
priority_rows = conn2.execute("""
    SELECT p.id, p.title, p.url, e.data_json
    FROM pages p
    JOIN extracted e ON p.id = e.page_id
    WHERE json_extract(e.data_json, '$.priority_tag') = 'YES'
    ORDER BY 
        CASE 
            WHEN json_extract(e.data_json, '$.budget_amount') IS NOT NULL 
             AND json_extract(e.data_json, '$.budget_amount') != 'null'
             AND json_extract(e.data_json, '$.budget_amount') != ''
            THEN 1 ELSE 2 END,
        json_extract(e.data_json, '$.budget_amount') DESC
""").fetchall()

p_idx = 0
for row in priority_rows:
    p_idx += 1
    data = json.loads(row['data_json'])
    
    values = [
        row['id'],
        row['title'],
        data.get('priority_reason', ''),
        data.get('budget_amount', ''),
        data.get('winning_price', ''),
        data.get('buyer', ''),
        data.get('winner', ''),
        data.get('region', ''),
        data.get('notice_type', ''),
        row['url'] or ''
    ]
    
    for col_idx, val in enumerate(values, 1):
        cell = ws2.cell(row=p_idx+1, column=col_idx, value=val)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        cell.fill = priority_fill
        if col_idx == 10 and val and val.startswith('http'):
            cell.font = url_font
            cell.hyperlink = val

ws2_col_widths = [6, 55, 25, 14, 16, 22, 22, 16, 14, 40]
for col, w in enumerate(ws2_col_widths, 1):
    ws2.column_dimensions[get_column_letter(col)].width = w
ws2.freeze_panes = 'A2'

# === 保存 ===
wb.save(output)

# 统计
priority_count = len(priority_rows)
normal_count = idx - priority_count

print("="*60)
print("📁 导出完成！")
print(f"  文件路径: {output}")
print(f"  总记录数: {idx}")
print(f"  重点项目: {priority_count}")
print(f"  一般项目: {normal_count}")
print(f"\n📄 Sheet 1: 完整数据 ({idx}条, 含自动筛选)")
print(f"📄 Sheet 2: 重点项目汇总 ({priority_count}条, 按预算排序)")
print("="*60)

conn2.close()
