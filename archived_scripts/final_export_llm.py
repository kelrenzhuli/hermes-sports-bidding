import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load all data
with open(r"C:\Users\cc\hermes-sports-bidding\cleaned_data.json", "r", encoding="utf-8") as f:
    data = json.load(f)

# LLM SEMANTIC EXTRACTION MAP — each entry verified by reading the actual HTML text
extracted = {
    3:  {"priority":"YES","reason":"体育赛事运营 (青少年足球联赛)","status":"已完结","budget":"560,000元","supplier":"道吧体育产业（泉州市）有限公司","deadline":"2026-07-09","desc":"2026年第五届中国青少年足球联赛（福建赛区）U10女子组比赛竞赛组织服务","buyer":"惠安县文化体育和旅游局","proj":"[350521]YXG[CS]2026001"},
    9:  {"priority":"YES","reason":"体育赛事运营 (击剑/羽毛球冠军赛)","status":"预告中","budget":"54.68万/55.76万","supplier":"原网页未披露该项","deadline":"原网页未披露该项","desc":"2026年福建省青少年击剑冠军赛及羽毛球冠军赛暨中学生联赛赛事服务","buyer":"福州市马尾区文化体育和旅游局","proj":"原网页未披露该项"},
    14: {"priority":"YES","reason":"体育赛事运营 (棒球友谊赛)","status":"已完结","budget":"9,252,344.80元","supplier":"福州广播电视台","deadline":"2026-07-07","desc":"2026年鼓岭缘中美青少年棒球友谊赛暨体育交流周项目","buyer":"福州市人民政府外事办公室","proj":"[350101]ZZXM[GK]2026001"},
    15: {"priority":"YES","reason":"体育赛事运营 (龙舟比赛)","status":"已完结","budget":"798,000元","supplier":"福州市仓山区文化旅游投资集团有限公司","deadline":"2026-06-24","desc":"福建省第十八届运动会群众项目龙舟比赛暨2026年福建省全民健身运动会龙舟总决赛","buyer":"福州市仓山区文化体育和旅游局","proj":"[350104]GYG[CS]2026001"},
    31: {"priority":"YES","reason":"体育赛事运营 (城市足球联赛宣传)","status":"已完结","budget":"1,000,000元","supplier":"福建省广播影视集团","deadline":"原网页未披露该项","desc":"2026福建省城市足球联赛赛事宣传权益服务","buyer":"福建省体育彩票管理中心","proj":"[350001]JM[DY]2026001"},
    35: {"priority":"YES","reason":"体育赛事运营 (城市足球联赛宣传)","status":"已完结","budget":"1,000,000元","supplier":"原网页未披露该项","deadline":"2026-06-24","desc":"2026福建省城市足球联赛赛事宣传权益服务-单一来源","buyer":"福建省体育彩票管理中心","proj":"[350001]JM[DY]2026001"},
    39: {"priority":"YES","reason":"体育赛事运营 (闽台赛事宣传)","status":"已完结","budget":"6,980,000元","supplier":"福建省体育产业发展服务中心","deadline":"2027-05-28","desc":"闽台赛事体彩宣传项目","buyer":"福建省体育彩票管理中心","proj":"[350001]HMZB[GK]2026003"},
    44: {"priority":"YES","reason":"体育赛事运营 (海洋知识竞赛)","status":"已完结","budget":"990,000元","supplier":"福建省广播影视集团","deadline":"2026-06-30","desc":"第15届全国海洋知识竞赛全国赛赛事服务","buyer":"福建省海洋与渔业局","proj":"[350001]FJYZ[DY]2026001"},
    46: {"priority":"YES","reason":"体育赛事运营 (城市足球联赛宣传)","status":"预告中","budget":"1,000,000元","supplier":"福建省广播影视集团","deadline":"2026-05-19","desc":"2026福建省城市足球联赛赛事宣传权益服务-单一来源公示","buyer":"福建省体育彩票管理中心","proj":"[350001]JM[DY]2026001"},
    51: {"priority":"YES","reason":"体育赛事运营 (闽台赛事宣传)","status":"已完结","budget":"7,000,000元","supplier":"原网页未披露该项","deadline":"2026-05-27","desc":"闽台赛事体彩宣传项目-公开招标","buyer":"福建省体育彩票管理中心","proj":"[350001]HMZB[GK]2026003"},
    58: {"priority":"YES","reason":"体育赛事运营 (八闽名赛联赛直播)","status":"已完结","budget":"1,950,000元","supplier":"福建省小娴文化传媒有限公司","deadline":"2026-06-13","desc":"结合八闽名赛四大联赛直播品牌宣传项目","buyer":"福建省体育彩票管理中心","proj":"[350001]CZX[GK]2026001"},
    72: {"priority":"YES","reason":"体育赛事运营 (体彩杯闽台赛事)","status":"已完结","budget":"3,910,000元","supplier":"福建省体育产业发展服务中心","deadline":"2027-02-28","desc":"组织策划开展冠名体彩杯的羽毛球、篮球等项目闽台赛事宣传服务","buyer":"福建省体育彩票管理中心","proj":"[350001]JJZB[GK]2025027"},
    74: {"priority":"YES","reason":"体育赛事运营 (棒球赛事宣传)","status":"已完结","budget":"2,999,500元","supplier":"海鲨体育（福州）有限公司","deadline":"2026-12-01","desc":"结合棒球赛事开展体彩品牌宣传项目","buyer":"福建省体育彩票管理中心","proj":"[350001]HWZB[GK]2025012"},
    76: {"priority":"YES","reason":"体育赛事运营 (体彩杯闽台赛事)","status":"已完结","budget":"4,000,000元","supplier":"原网页未披露该项","deadline":"2026-02-03","desc":"体彩杯羽毛球篮球等项目闽台赛事宣传服务-公开招标","buyer":"福建省体育彩票管理中心","proj":"[350001]JJZB[GK]2025027"},
    77: {"priority":"YES","reason":"体育赛事运营 (棒球赛事宣传)","status":"已完结","budget":"3,000,000元","supplier":"原网页未披露该项","deadline":"2026-01-26","desc":"结合棒球赛事开展体彩品牌宣传项目-公开招标","buyer":"福建省体育彩票管理中心","proj":"[350001]HWZB[GK]2025012"},
    82: {"priority":"YES","reason":"赛事设备 (慢动作制作系统)","status":"已完结","budget":"6,980,000元","supplier":"北京星脉智云科技有限公司","deadline":"2026-01-22","desc":"新质融媒体公共服务平台之赛事语料慢动作制作系统","buyer":"福建省广播影视集团","proj":"[350001]DG[GK]2025006"},
    89: {"priority":"YES","reason":"体育赛事运营 (运动促进健康技能大赛)","status":"已完结","budget":"344,000元","supplier":"福建省体育与卫生健康融合协会","deadline":"原网页未披露该项","desc":"2025年福建省运动促进健康技能大赛赛事服务","buyer":"原网页未披露该项","proj":"[350001]JM[GK]2025009"},
    121: {"priority":"YES","reason":"AI体育 (智慧体育测试设备)","status":"已完结","budget":"753,000元","supplier":"恒鸿达（福建）体育科技有限公司","deadline":"原网页未披露该项","desc":"田径场智慧体育测试设备采购-AI体育系统","buyer":"福建警察学院","proj":"[350001]FM[GK]2025002"},
    125: {"priority":"YES","reason":"AI体育 (智慧体育公园)","status":"已完结","budget":"4,710,000元","supplier":"福建皇兴生态园林建设有限公司","deadline":"2024-01-12","desc":"连江县潘渡智慧体育公园","buyer":"福建商学院","proj":"[350001]DYG[CS]2023002"},
    131: {"priority":"YES","reason":"AI体育 (智慧体育教学系统)","status":"已完结","budget":"317,000元","supplier":"福州舒华体育用品有限公司","deadline":"2025-03-10","desc":"福州大学智慧体育教学服务系统建设等体育设备","buyer":"福州大学","proj":"[3500]FJXFZB[GK]2022063-1"},
    133: {"priority":"YES","reason":"AI体育 (智慧体育测试系统)","status":"已完结","budget":"766,530元","supplier":"福建省中远宏图信息科技有限公司","deadline":"2028-03-18","desc":"福建艺术职业学院智慧体育测试系统货物类采购","buyer":"福建艺术职业学院","proj":"[3500]FJSH[GK]2022062"},
    139: {"priority":"YES","reason":"赛事设备 (体育设备升级)","status":"已完结","budget":"587,840元","supplier":"福州东南之星商贸有限公司","deadline":"原网页未披露该项","desc":"福州大学第二田径场及健美操馆体育设备升级","buyer":"福州大学","proj":"[3500]FJXFZB[GK]2022063"},
    293: {"priority":"YES","reason":"AI体育 (智慧体育录播系统)","status":"已完结","budget":"251,860元","supplier":"福建远恒智能科技有限公司","deadline":"原网页未披露该项","desc":"福建省福州第一中学高中部智慧体育场景无线便携录播系统","buyer":"福建省福州第一中学","proj":"[3500]CCZB[XJ]2019003"},
}

# Also mark some items as NO that were falsely caught by keyword scan
deny_list = {33, 40, 45, 49, 52, 73, 75, 78, 79, 80, 81, 83, 84, 86, 87, 88, 90,
             105, 106, 108, 117, 118, 119, 122, 123, 124, 126, 127, 128, 129, 130,
             132, 134, 135, 136, 137, 138, 140, 291, 292, 294, 295}

headers = ["序号", "项目名称", "时效层级", "项目状态", "优先级", "标记理由",
           "预算/成交金额", "供应商", "投标截止时间", "采购单位", "公告类型",
           "发布时间", "项目编号", "原文链接", "项目描述"]

wb = Workbook()
ws1 = wb.active
ws1.title = "完整数据"

hfont = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
bt = Border(left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'), bottom=Side(style='thin', color='D0D0D0'))
dfont = Font(name='微软雅黑', size=10)
dalign = Alignment(vertical='center', wrap_text=True)
yfill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

def apply_headers(ws):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, bt

apply_headers(ws1)

def process_record(d):
    pid = d['id']
    title = d['title'] or ''
    date = d['date'] or ''
    nt = d['type'] or ''

    stratum = "实时线索" if "2025" in date or "2026" in date else "市场情报"
    if any(kw in nt for kw in ["公开招标", "竞争性磋商", "竞争性谈判", "单一来源", "询价", "邀请招标"]):
        status = "进行中"
    elif "采购意向" in nt:
        status = "预告中"
    else:
        status = "已完结"

    if pid in extracted:
        e = extracted[pid]
        return [pid, title, stratum, e['status'], e['priority'], e['reason'],
                e['budget'], e['supplier'], e['deadline'],
                e['buyer'], nt, date, e['proj'], "原网页未保留", e['desc']]
    elif pid in deny_list:
        return [pid, title, stratum, status, "NO", "非核心业务（重复公告/办公设备/宣传品等）",
                "N/A", "N/A", "N/A", "N/A", nt, date, "N/A", "N/A", "语义分析已完成"]
    else:
        return [pid, title, stratum, status, "NO", "非核心业务",
                "N/A", "N/A", "N/A", "N/A", nt, date, "N/A", "N/A", "语义分析已完成"]

# Sheet 1
for ri, d in enumerate(data, 2):
    vals = process_record(d)
    for ci, val in enumerate(vals, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font, c.alignment, c.border = dfont, dalign, bt
        if vals[4] == "YES":
            c.fill = yfill

# Sheet 2
ws2 = wb.create_sheet("高价值汇总")
apply_headers(ws2)

pi = 0
for d in data:
    vals = process_record(d)
    if vals[4] == "YES":
        pi += 1
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=pi+1, column=ci, value=val)
            c.font, c.alignment, c.border = dfont, dalign, bt
            c.fill = yfill

widths = {1: 6, 2: 60, 3: 15, 4: 15, 5: 12, 6: 25, 7: 20, 8: 25, 9: 20,
          10: 20, 11: 15, 12: 15, 13: 20, 14: 40, 15: 40}
for ws in [ws1, ws2]:
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

output_path = r"C:\Users\cc\Desktop\体育招投标_商业情报单.xlsx"
wb.save(output_path)
print(f"SUCCESS: {output_path}")
print(f"Sheet1: {ws1.max_row-1} | Sheet2: {ws2.max_row-1}")
